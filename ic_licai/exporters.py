from typing import Dict, Any, List
from fpdf import FPDF
from openpyxl import Workbook
import io, json, datetime

PRIMARY = (10, 38, 64)  # #0A2640

class PDF(FPDF):
    def _bullet(self: text: str, indent: int = 10):
    """
    Safe indented line: moves cursor to left margin + indent and
    uses an explicit width that always fits the page.
    """
    # Move to a safe X position: left margin + indent
    self.set_x(self.l_margin + indent)
    usable_w = self.w - self.r_margin - self.get_x()
    if usable_w < 20:
       self.set_x(self.l_margin)
    self.multi_cell(usable_w, 6, text)
    def header(self):
        self.set_text_color(*PRIMARY)
        self.set_font("Helvetica", "B", 16)
        self.set_xy(15, 12)
        self.cell(0, 10, getattr(self, "header_title", ""), ln=1)
        self.set_text_color(0, 0, 0)
        self.set_font("Helvetica", "", 9)
        self.set_xy(15, 22)
        self.cell(0, 6, f"Generated {datetime.datetime.utcnow().isoformat()} (UTC) - IC-LicAI Demo", ln=1)
        self.set_draw_color(*PRIMARY)
        self.set_line_width(0.4)
        self.line(15, 30, 195, 30)
        self.ln(8)

def _wrap_text(pdf: FPDF, text: str, width: int=0):
    for line in text.split("\n"):
        pdf.multi_cell(0 if width==0 else width, 6, line)

def export_pdf(data: Dict[str, Any]) -> bytes:
    pdf = PDF(format="A4")
    pdf.set_auto_page_break(auto=True, margin=18)

    # Cover / summary
    pdf.add_page(); pdf.header_title = "Intangible Capital & Licensing Readiness Report"
    pdf.set_font("Helvetica", "", 12); pdf.ln(20)
    pdf.cell(0, 8, f"Case: {data.get('case','(unspecified)')}", ln=1)
    pdf.ln(2); pdf.set_font("Helvetica", "B", 12); pdf.cell(0, 8, "Executive Summary", ln=1)
    pdf.set_font("Helvetica", "", 10)
    _wrap_text(pdf, data.get("summary", "This advisory report summarises the organisation's intangible capital and licensing readiness."))

    # IC map
    pdf.add_page(); pdf.header_title = "Intangible Capital Map (4-Leaf)"
    pdf.set_font("Helvetica", "", 10)
    for leaf, items in (data.get("ic_map") or {}).items():
        pdf.set_font("Helvetica", "B", 11); pdf.cell(0, 7, f"* {leaf}", ln=1)
        pdf.set_font("Helvetica", "", 10)
        for it in items[:6]:
            _bullet(pdf, f"- {it}")
        pdf.ln(2)

    # 10 Steps readiness
    pdf.add_page(); pdf.header_title = "10-Steps Readiness Summary"
    for row in data.get("readiness", []):
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(0, 6, f"Step {row['step']}: {row['name']} (score {row['score']}/3)", ln=1)
        pdf.set_font("Helvetica", "", 10)
        for t in row["tasks"]:
            _bullet(pdf, f"- {t}")
        pdf.ln(2)

    # Licensing options
    pdf.add_page(); pdf.header_title = "Licensing Options (advisory)"
    for opt in data.get("licensing", []):
        pdf.set_font("Helvetica", "B", 11); pdf.cell(0, 7, f"* {opt['model']}", ln=1)
        pdf.set_font("Helvetica", "", 10)
        _bullet(pdf, opt["notes"])
        pdf.ln(1)

    # Governance note
    pdf.add_page(); pdf.header_title = "Governance & Audit Note"
    pdf.set_font("Helvetica", "", 10)
    _wrap_text(pdf, "This report is generated using an advisory-first workflow with human approval. Evidence sources and decisions are auditable.")

    out = io.BytesIO(); pdf.output(out); out.seek(0); return out.getvalue()

def export_xlsx(ic_map: Dict[str, List[str]]) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "IA Register"
    ws.append(["Asset Name","Capital Type","Tacit/Explicit","Evidence Source","IAS 38 Status","Priority","Next Step"])
    for leaf, items in ic_map.items():
        for it in items:
            ws.append([it, leaf, "", "", "", "", ""])
    bio = io.BytesIO(); wb.save(bio); bio.seek(0); return bio.getvalue()

def export_json(data: Dict[str, Any]) -> bytes:
    return json.dumps(data, indent=2).encode("utf-8")
